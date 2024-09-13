"""

This is a command line interface for the Easy Access project.
It can be used to:
-> Read data from a CopyRight export
-> Process the data into a standard format for usage by people who check the documents on canvas
-> Export various sheets:
    -> per faculty
    -> all items
    -> only items that have been changed
    -> etc
-> Read back the data from all the sheets to produce a CopyRight 'import' sheet

"""

import openpyxl.worksheet
import openpyxl.worksheet.datavalidation
import typer
from typing_extensions import Annotated
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
import polars as pl
from pathlib import Path
import os
from datetime import datetime
import dotenv
from enum import Enum
import json
import openpyxl

from file_utils import Directory, File

print: callable = Console(emoji=True, markup=True).print
dotenv.load_dotenv('settings.env')

def info(text: str):
    print(f":information: [cyan] {text} [/cyan]")

def warn(text: str):
    print(f":warning: [bold red] {text} [/bold red]")

def cool(text: str):
    print(f":smiling_face_with_sunglasses: [yellow] {text} [/yellow]")

class Functions(str, Enum):
    both = "both"
    read = "read"
    export = "export"


# The main CLI function:
def cli(do: Annotated[Functions, typer.Option(case_sensitive=False, help="Which tool to run: read in new data, export current data, or both.", rich_help_panel="Functions")] = Functions.read,
        changes: Annotated[bool, typer.Option(help="Only add items that have been changed to new faculty sheets.", rich_help_panel="Functions")] = True,
        copyright_export_dir: Annotated[str | None, typer.Argument(help="Overwrite location for files exported from CopyRight", rich_help_panel="Overwite Directory Locations")] = None,
        copyright_import_dir: Annotated[str | None, typer.Argument(help="Overwrite location for files to be imported back into CopyRight", rich_help_panel="Overwite Directory Locations")] = None,
        faculties_dir: Annotated[str | None, typer.Argument(help="Overwrite location for faculty sheets", rich_help_panel="Overwite Directory Locations")] = None,
        all_items_dir: Annotated[str | None, typer.Argument(help="Overwrite location for 'all items' sheet", rich_help_panel="Overwite Directory Locations")] = None,
        ):
    """
    Runs the Easy Access toolkit with the specified settings.\n
    Make sure that these two files are present in the current dir and contain the required info:\n\n
        'settings.env': The directories to use\n
        'department_mapping.json': The mapping between department names and faculty names\n
    \n
    Visit the repo for more instructions & the latest version: https://github.com/utsmok/ea-cli. (<- you can click this in your terminal!)\n
    \n\n
    Example usage\n
    --------------\n
    ea-cli\n
    ea-cli --do export\n
    ea-cli --no-changes\n
    ea-cli --do read --changes\n
    ea-cli --do both --copyright_export_dir 'C:/easy_access_sheets/cli_copyright_data'  \n
    """

    if do not in [Functions.both, Functions.read, Functions.export]:
        warn("No functions selected! Aborting the script. Next time, enable at least one of 'Function' options; for details run ea-cli --help.")
        cool("Thank you for using the Easy Access tool!")
        raise typer.Exit(code=1)


    if all(not x for x in [copyright_export_dir, copyright_import_dir, faculties_dir, all_items_dir]):
        dirs = None
    else:
        dirs={
            'copyright_export':copyright_export_dir,
            'copyright_import':copyright_import_dir,
            'faculties':faculties_dir,
            'all_items':all_items_dir,
        }

    tool = EasyAccessTool(functions=do, only_changes=changes, dirs=dirs)
    tool.run()

    cool("All done! Thank you for using the Easy Access tool!")


class EasyAccessTool:

    settings: list[callable] = []
    files: dict[str,File]
    dirs: dict[str,Directory] = {
        'root':Directory(os.getcwd()),
        'copyright_export':Directory(os.getenv("COPYRIGHT_EXPORT_DIR")),
        'copyright_import':Directory(os.getenv("COPYRIGHT_IMPORT_DIR")),
        'all_items':Directory(os.getenv("ALL_ITEMS_DIR")),
        'faculties':Directory(os.getenv("FACULTIES_DIR")),
    }

    raw_copyright_data: pl.DataFrame
    copyright_data: pl.DataFrame
    faculty_sheet_data: pl.DataFrame
    all_items_sheet_data: pl.DataFrame
    dept_mapping_path = File("department_mapping.json")
    DEPARTMENT_MAPPING = json.load(open(dept_mapping_path.path, encoding='utf-8'))
    faculties: list[str]
    latest_file: File
    latest_file_date: str

    def __init__(self, functions: Functions | None = Functions.both, dirs: dict[str,str] | None = None, only_changes: bool = True) -> None:
        """
        Parameters:
            setting:  str | None
                Pick which functions to run when self.run() is called. If no argument is passed, it will run all functions.
                pick from one of the presets below:
                    'none' -> don't run any functions
                    'all' -> run all functions
                    'new_data' -> read in new data, process it, create new faculty and all itemssheets
                    'read_sheets' -> read in faculty sheet data, process, create import sheet
            dirs: dict[str,str] | None
                A dict containing the str path to the directories to use. If not provided, it will use the default dirs from settings.env.
            only_changes: bool = True
                True (default): only add items that have been changed to the created sheets
                False: add all items from the CopyRight export to the created sheets
        """
        ...


        if functions is None:
            self.settings = []
        elif functions == Functions.both:
            self.settings = [self.read_copyright_export, self.process_copyright_export,  # read in new data
                        self.read_all_items_sheet, self.read_faculty_sheets, # read in data manually added to sheets
                        self.create_import_sheet,  # from the old data, create a sheet to import into CopyRight
                        self.create_faculty_sheets, self.create_all_items_sheet] # create new sheets with new data
        elif functions == Functions.read:
            self.settings = [self.read_copyright_export, self.process_copyright_export,  # read in new data
                        self.create_faculty_sheets, self.create_all_items_sheet] # create new sheets with new data
        elif functions == Functions.export:
            self.settings = [self.read_faculty_sheets,  # read in data manually added to sheets
                        self.create_import_sheet,  # from the old data, create a sheet to import into CopyRight
                        ]
        if dirs:
            for key, value in dirs.items():
                if value:
                    self.dirs[key] = Directory(value)

        self.only_changes = only_changes


    def run(self) -> None:
        """
        Runs the functions as specified in the settings dict.
        """
        for func in self.settings:
            func()


    def read_copyright_export(self) -> None:
        ...
        # determine which file to grab
        # then read it in
        # store it in self.raw_copyright_data
        info(f"Reading in newest Copyright Data from {self.dirs['copyright_export']}")
        try:
            all_files = self.dirs['copyright_export'].files
            self.latest_file = max(all_files, key=lambda x: x.created)
            self.latest_file_date = self.latest_file.created.strftime("%Y-%m-%d")
            info(f"Selected newest file: {self.latest_file}")
            self.raw_copyright_data = pl.read_excel(self.latest_file.path)
        except FileNotFoundError:
            warn(f"No files found in {self.dirs['copyright_export']}")
            raise typer.Exit(code=1)
        except PermissionError:
            warn(f"Permission denied to read {self.latest_file}")
            raise typer.Exit(code=1)
        except ValueError:
            warn("No files found in {self.dirs['copyright_export']}")
            raise typer.Exit(code=1)

    def process_copyright_export(self) -> None:
        # process self.raw_copyright_data
        # change the format to match the UT EA format
        # if only_changes: read in faculty sheets to create a list of all items, compare with self.raw_copyright_data
        # skip each item with identical material_id and last_change values, keep the rest for this export
        # else: don't compare and just keep the entire dataframe
        # store the result in self.copyright_data

        self.copyright_data = self.raw_copyright_data.rename(
            lambda col: col.replace(" ", "_")
            .replace("#", "count_")
            .replace("*", "x")
            .lower()
            ).with_columns(
            pl.Series("retrieved_from_copyright_on", [self.latest_file_date] * len(self.raw_copyright_data)),
            pl.Series("workflow_status", ["ToDo"] * len(self.raw_copyright_data)),
            faculty=pl.col("department").replace_strict(
                self.DEPARTMENT_MAPPING, default="Unmapped"
            ),
        )

        if self.only_changes:
            #self.read_faculty_sheets()
            self.faculty_sheet_data = pl.DataFrame()
            if self.faculty_sheet_data.is_empty():
                info("No faculty sheets found. Adding all items without checking for changes.")
            else:
                # compare self.copyright_data with self.faculty_sheet_data
                # only keep items from self.copyright_data with a value in col material_id that is not in self.faculty_sheet_data
                # AND items with a matching material_id but a different value in col last_change
                not_in_faculty = self.copyright_data.join(
                    self.faculty_sheet_data,
                    on="material_id",
                    how="anti"
                )
                matching_id_diff_change = self.copyright_data.join(
                    self.faculty_sheet_data,
                    on="material_id",
                    how="inner"
                ).filter(
                    pl.col("last_change") != pl.col("last_change_right")
                ).select(
                    pl.all().exclude("last_change_right")
                ).drop_nulls(
                    pl.col("material_id")
                ).filter(
                    pl.col("status") == "Deleted"
                )
                self.copyright_data = pl.concat([not_in_faculty, matching_id_diff_change])


        self.faculties = self.copyright_data.select(pl.col("faculty").unique()).to_series().to_list()

    def create_faculty_sheets(self) -> None:
        # from self.copyright_data, create a sheet for each faculty (see col 'faculty')
        # store the excel files in self.dirs['faculties']
        for faculty in self.faculties:
            faculty_dir = Directory(self.dirs['faculties'].full / faculty)
            if faculty is None or faculty == "":
                faculty = "no_faculty_found"
            filename = f"{faculty}_{self.latest_file_date}.xlsx"
            i = 1
            while os.path.exists(faculty_dir.full / filename):
                filename = f"{faculty}_{self.latest_file_date}_{i}.xlsx"
                i += 1

            faculty_data = self.copyright_data.filter(pl.col("faculty") == faculty)

            if faculty_data.is_empty():
                warn(f"No items found for faculty {faculty}.")
            #TODO: create individual sheets within the excel file, for now just write the whole thing to one sheet

            faculty_data.write_excel(faculty_dir.full / filename)
            info(f"Created sheet: {faculty_dir.full / filename}")
            self.finalize_sheet(File(str(faculty_dir.full / filename)))


    def finalize_sheet(self, file: File) -> None:
        """
        For a given freshly created excel file,
        split it into multiple sheets, and add dropdowns for certain columns.
        """

        # load excel
        wb = openpyxl.load_workbook(filename = str(file.path))

        # add sheet + names
        wb.active.title = 'Complete data'
        entry_sheet = wb.create_sheet('Data entry')


        keep_cols = [6, 34, 14, 16, 17, 13, 1, 8, 9, 28, 3, 5]
        col_names = ['url', 'workflow_status', 'manual_classification', 'scope', 'remarks', 'ml_prediction',
                'material_id', 'title', 'owner', 'author', 'department', 'course_name']

        for new_col, old in enumerate(keep_cols, start=1):
            for row, cell in enumerate(wb.active.iter_rows(min_col=old, max_col=old, values_only=True), start=1):
                entry_sheet.cell(row=row, column=new_col).value = cell[0]

        # add dropdowns
        for col, name in enumerate(col_names, start=1):
            entry_sheet.cell(row=1, column=col, value=name)


        dropdowndata = [(2,'B', '"ToDo,Done,InProgress"'), # workflow status
                        (3,'C', '"open access, eigen materiaal - powerpoint, eigen materiaal - overig, lange overname, eigen materiaal - titelindicatie"'), # manual classification
                        ]
        for colnum, col_letter, itemlist in dropdowndata:
            dv = openpyxl.worksheet.datavalidation.DataValidation(type="list", formula1=itemlist, allow_blank=False)
            dv.error = "Please select a valid option from the list"
            dv.errorTitle = "Invalid option"
            dv.prompt= "Please select from the list"
            dv.promptTitle = "List selection"
            entry_sheet.add_data_validation(dv)
            dv.add(f"{col_letter}1:{col_letter}1000")

        wb.save(filename = str(file.path))





    def create_all_items_sheet(self) -> None:
        # from self.copyright_data, create a sheet with all items
        # store the excel file in self.dirs['all_items']
        filename = f"all_items_{self.latest_file_date}.xlsx"
        i = 1
        while os.path.exists(self.dirs['all_items'].full / filename):
            filename = f"all_items_{self.latest_file_date}_{i}.xlsx"
            i += 1

        self.copyright_data.write_excel(self.dirs['all_items'].full / filename)
        info(f"Created sheet: {self.dirs['all_items'].full} / {filename}")
    def read_faculty_sheets(self) -> None:
        # read in all the faculty sheets in self.dirs['faculties']
        # combine and store in self.faculty_sheet_data
        self.faculty_sheet_data = self.read_sheets(self.dirs['faculties'].files_r)

    def read_all_items_sheet(self) -> None:
        # read in 'all items' sheet(s) in self.dirs['all_items']
        # combine and store in self.all_items_sheet_data
        self.all_items_sheet_data  = self.read_sheets(self.dirs['all_items'].files_r)

    def read_sheets(self, files: list[File]) -> pl.DataFrame:
        # TODO: handle multiple sheets in the same file
        file_data = []
        for file in files:
            current_data = pl.read_excel(file.path)
            current_data = self.validate_ea_sheet(current_data, file)
            if current_data.is_empty():
                continue
            else:
                file_data.append(current_data)
        if file_data:
            result: pl.DataFrame = pl.concat(file_data)
        else:
            result = pl.DataFrame()
        return result.unique()

    def validate_ea_sheet(self, df: pl.DataFrame, file:File) -> pl.DataFrame:
        """
        For a given dataframe created from an EA excel sheet,
        check that the data is in the correct format, and the values are correct.
        If not, try to fix, else print the errors.
        If the sheet is not validated, return an empty dataframe.
        """
        valid = True
        errlist = []
        # TODO: handle multiple sheets in the same file
        warn(':triangular_flag: Sheet validation is not implemented at the moment.')
        if df.is_empty():
            valid = False
            errlist.append("Sheet is empty")

        if valid:
            # check that the sheet has the correct columns
            ...
        if valid:
            # check the values in the columns
            ...



        if not valid:
            info(f"Errors in sheet {file}:")
            for err in errlist:
                warn(err)
            return pl.DataFrame()
        else:
            return df

    def create_import_sheet(self) -> None:
        '''
        combine self.faculty_sheet_data and self.all_items_sheet_data
        clean it up
        change from UT Easy Access format to SURF CopyRight format
        create & export an .xlsx sheet that can be sent to SURF to be imported into CopyRight.
        '''
        ...

if __name__ == "__main__":
    typer.run(cli)